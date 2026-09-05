"""
routes.py — API REST per la gestione delle ricevute sanitarie.

Endpoints:
  POST   /api/ricevute           → Crea nuova ricevuta (calcola progressivo annuale)
  GET    /api/ricevute            → Lista ricevute (con filtri opzionali anno/paziente)
  GET    /api/ricevute/{id}       → Dettaglio singola ricevuta (per ristampa)
  GET    /api/ricevute/prossimo-numero?anno=  → Prossimo numero progressivo
  GET    /api/pazienti/cerca?cf=  → Cerca paziente per Codice Fiscale (autocomplete)
  GET    /api/ricevute/esporta-csv?anno=  → Esporta CSV annuale
"""

import csv
import io
import os
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy import func, extract
from sqlalchemy.orm import Session, joinedload

from database import get_db
from models import Paziente, Ricevuta, RigaRicevuta
from schemas import (
    PazienteOut,
    RicevutaIn,
    RicevutaOut,
    RicevutaListItem,
)

router = APIRouter(prefix="/api", tags=["ricevute"])


# ---------------------------------------------------------------------------
# Utilità interna: calcolo prossimo numero progressivo annuale
# ---------------------------------------------------------------------------
def _prossimo_numero(db: Session, anno: int) -> int:
    """
    Trova il primo numero progressivo disponibile a partire da 1 per l'anno specificato.
    Se sono stati eliminati numeri o ci sono buchi nella numerazione, riutilizza il primo buco.
    """
    numeri_usati = set(
        row[0] for row in db.query(Ricevuta.numero).filter(Ricevuta.anno == anno).all()
    )
    candidato = 1
    while candidato in numeri_usati:
        candidato += 1
    return candidato


# ---------------------------------------------------------------------------
# POST /api/ricevute — Crea nuova ricevuta
# ---------------------------------------------------------------------------
@router.post("/ricevute", response_model=RicevutaOut, status_code=201)
def crea_ricevuta(payload: RicevutaIn, db: Session = Depends(get_db)):
    """
    Crea una nuova ricevuta sanitaria.

    - Cerca il paziente per Codice Fiscale; se non esiste, lo crea.
    - Calcola il numero progressivo annuale.
    - Calcola subtotale, bollo e totale da pagare.
    """
    # 1. Cerca o crea il paziente
    paziente = (
        db.query(Paziente)
        .filter(Paziente.codice_fiscale == payload.paziente.codice_fiscale)
        .first()
    )
    if paziente is None:
        paziente = Paziente(**payload.paziente.model_dump())
        db.add(paziente)
        db.flush()  # Ottiene l'ID senza committare
    else:
        # Aggiorna i dati anagrafici con quelli più recenti
        for campo, valore in payload.paziente.model_dump().items():
            setattr(paziente, campo, valore)
        db.flush()

    # 2. Calcola o valida il numero progressivo
    anno = payload.data_emissione.year
    if payload.numero is not None:
        esistente = (
            db.query(Ricevuta)
            .filter(Ricevuta.anno == anno, Ricevuta.numero == payload.numero)
            .first()
        )
        if esistente is not None:
            raise HTTPException(
                status_code=400,
                detail=f"La ricevuta n. {anno}/{payload.numero} esiste già nel database",
            )
        numero = payload.numero
    else:
        numero = _prossimo_numero(db, anno)

    # 3. Calcola i totali
    subtotale = sum(riga.importo for riga in payload.righe)

    # Logica bollo: se subtotale > 77,47 € e bollo richiesto
    bollo_applicato = payload.bollo_applicato
    importo_bollo = Decimal("2.00") if bollo_applicato else Decimal("0.00")

    totale_da_pagare = subtotale
    if bollo_applicato and payload.bollo_a_carico_paziente:
        totale_da_pagare += importo_bollo

    # 4. Crea la ricevuta
    ricevuta = Ricevuta(
        anno=anno,
        numero=numero,
        data_emissione=payload.data_emissione,
        paziente_id=paziente.id,
        subtotale=subtotale,
        bollo_applicato=bollo_applicato,
        bollo_a_carico_paziente=payload.bollo_a_carico_paziente,
        importo_bollo=importo_bollo,
        totale_da_pagare=totale_da_pagare,
        metodo_pagamento=payload.metodo_pagamento,
        note=payload.note,
    )
    db.add(ricevuta)
    db.flush()

    # 5. Crea le righe di prestazione
    for riga_in in payload.righe:
        riga = RigaRicevuta(
            ricevuta_id=ricevuta.id,
            ordine=riga_in.ordine,
            descrizione=riga_in.descrizione,
            importo=riga_in.importo,
        )
        db.add(riga)

    db.commit()
    db.refresh(ricevuta)

    return ricevuta


# ---------------------------------------------------------------------------
# GET /api/ricevute — Lista ricevute (archivio)
# ---------------------------------------------------------------------------
@router.get("/ricevute", response_model=list[RicevutaListItem])
def lista_ricevute(
    anno: Optional[int] = Query(None, description="Filtra per anno"),
    paziente: Optional[str] = Query(None, description="Filtra per cognome, nome o CF"),
    db: Session = Depends(get_db),
):
    """
    Restituisce la lista delle ricevute emesse, ordinata per anno DESC e numero DESC.
    Filtri opzionali per anno e ricerca paziente (cognome, nome o CF).
    """
    query = (
        db.query(
            Ricevuta.id,
            Ricevuta.anno,
            Ricevuta.numero,
            Ricevuta.data_emissione,
            Paziente.cognome.label("paziente_cognome"),
            Paziente.nome.label("paziente_nome"),
            Paziente.codice_fiscale.label("codice_fiscale"),
            Ricevuta.subtotale,
            Ricevuta.totale_da_pagare,
            Ricevuta.metodo_pagamento,
        )
        .join(Paziente, Ricevuta.paziente_id == Paziente.id)
    )

    if anno is not None:
        query = query.filter(Ricevuta.anno == anno)

    if paziente is not None and paziente.strip():
        term = f"%{paziente.strip()}%"
        from sqlalchemy import or_
        query = query.filter(
            or_(
                Paziente.cognome.ilike(term),
                Paziente.nome.ilike(term),
                Paziente.codice_fiscale.ilike(term),
            )
        )

    rows = query.order_by(Ricevuta.anno.desc(), Ricevuta.numero.desc()).all()

    return [
        RicevutaListItem(
            id=r.id,
            anno=r.anno,
            numero=r.numero,
            data_emissione=r.data_emissione,
            paziente_cognome=r.paziente_cognome,
            paziente_nome=r.paziente_nome,
            codice_fiscale=r.codice_fiscale,
            subtotale=r.subtotale,
            totale_da_pagare=r.totale_da_pagare,
            metodo_pagamento=r.metodo_pagamento,
        )
        for r in rows
    ]



# ---------------------------------------------------------------------------
# GET /api/ricevute/prossimo-numero — Prossimo numero progressivo
# (DEVE stare prima di /ricevute/{ricevuta_id} per evitare conflitto di routing)
# ---------------------------------------------------------------------------
@router.get("/ricevute/prossimo-numero")
def prossimo_numero(
    anno: int = Query(..., description="Anno di riferimento"),
    db: Session = Depends(get_db),
):
    """Restituisce il prossimo numero progressivo disponibile per l'anno indicato."""
    return {"anno": anno, "prossimo_numero": _prossimo_numero(db, anno)}


# ---------------------------------------------------------------------------
# GET /api/ricevute/verifica-numero — Verifica disponibilità numero
# (DEVE stare prima di /ricevute/{ricevuta_id})
# ---------------------------------------------------------------------------
@router.get("/ricevute/verifica-numero")
def verifica_numero(
    anno: int = Query(..., description="Anno"),
    numero: int = Query(..., ge=1, description="Numero da verificare"),
    escludi_id: Optional[int] = Query(None, description="ID ricevuta da escludere (in modifica)"),
    db: Session = Depends(get_db),
):
    """Verifica se la combinazione (anno, numero) è già utilizzata."""
    query = db.query(Ricevuta).filter(Ricevuta.anno == anno, Ricevuta.numero == numero)
    if escludi_id is not None:
        query = query.filter(Ricevuta.id != escludi_id)
    esiste = query.first() is not None
    return {"disponibile": not esiste}


# ---------------------------------------------------------------------------
# GET /api/ricevute/anni — Elenco anni con ricevute presenti
# (DEVE stare prima di /ricevute/{ricevuta_id})
# ---------------------------------------------------------------------------
@router.get("/ricevute/anni")
def elenco_anni(db: Session = Depends(get_db)):
    """Restituisce l'elenco di tutti gli anni distinti in cui ci sono ricevute emesse."""
    anni_db = db.query(Ricevuta.anno).distinct().order_by(Ricevuta.anno.desc()).all()
    anni = [r[0] for r in anni_db]
    anno_corrente = date.today().year
    if anno_corrente not in anni:
        anni.append(anno_corrente)
        anni.sort(reverse=True)
    return {"anni": anni}


# ---------------------------------------------------------------------------
# GET /api/ricevute/esporta-csv — Export CSV annuale
# (DEVE stare prima di /ricevute/{ricevuta_id} per evitare conflitto di routing)
# ---------------------------------------------------------------------------
@router.get("/ricevute/esporta-csv")
def esporta_csv(
    anno: Optional[int] = Query(None, description="Anno solare da esportare"),
    data_inizio: Optional[date] = Query(None, description="Data inizio intervallo (inclusa)"),
    data_fine: Optional[date] = Query(None, description="Data fine intervallo (inclusa)"),
    db: Session = Depends(get_db),
):
    """
    Genera un file CSV con il riepilogo contabile delle prestazioni emesse nel periodo specificato,
    strutturato con le colonne richieste dalla contabilità:
    DATA | COGNOME NOME | VIA CITTA' CAP | CODICE FISCALE | PRESTAZIONE | IMPORTO
    """
    query = (
        db.query(Ricevuta)
        .options(joinedload(Ricevuta.paziente), joinedload(Ricevuta.righe))
    )

    if data_inizio is not None:
        query = query.filter(Ricevuta.data_emissione >= data_inizio)
    if data_fine is not None:
        query = query.filter(Ricevuta.data_emissione <= data_fine)
    if anno is not None and data_inizio is None and data_fine is None:
        query = query.filter(Ricevuta.anno == anno)

    ricevute = query.order_by(Ricevuta.data_emissione.asc(), Ricevuta.numero.asc()).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    headers = [
        "DATA",
        "COGNOME NOME",
        "VIA CITTA' CAP",
        "CODICE FISCALE",
        "PRESTAZIONE",
        "IMPORTO",
    ]
    writer.writerow(headers)

    for r in ricevute:
        paz = r.paziente
        cognome_nome = f"{paz.cognome or ''} {paz.nome or ''}".strip().upper()

        # Composizione VIA CITTA' CAP
        parti_ind = []
        if paz.indirizzo and paz.indirizzo.strip():
            parti_ind.append(paz.indirizzo.strip())
        if paz.citta and paz.citta.strip():
            parti_ind.append(paz.citta.strip())
        if paz.cap and paz.cap.strip():
            parti_ind.append(paz.cap.strip())
        via_citta_cap = " ".join(parti_ind).upper() if parti_ind else ""

        data_str = r.data_emissione.strftime("%d/%m/%Y")
        cf_str = (paz.codice_fiscale or "").strip().upper()

        # Tutte le prestazioni della fattura in fila (separate da virgola)
        descrizioni = [riga.descrizione.strip() for riga in r.righe if riga.descrizione and riga.descrizione.strip()]
        prestazioni_str = ", ".join(descrizioni)

        # Importo totale complessivo della fattura (senza riga della marca da bollo)
        totale_val = f"{float(r.totale_da_pagare or 0):.2f}".replace(".", ",")

        writer.writerow([
            data_str,
            cognome_nome,
            via_citta_cap,
            cf_str,
            prestazioni_str,
            totale_val,
        ])

    filename_parts = ["ricevute"]
    if anno and not data_inizio and not data_fine:
        filename_parts.append(str(anno))
    else:
        if data_inizio:
            filename_parts.append(f"da_{data_inizio.strftime('%Y%m%d')}")
        if data_fine:
            filename_parts.append(f"a_{data_fine.strftime('%Y%m%d')}")
    filename = f"{'_'.join(filename_parts)}.csv"

    csv_content = output.getvalue().encode("utf-8-sig")

    saved_path_str = ""
    try:
        downloads_dir = Path.home() / "Downloads"
        if downloads_dir.exists():
            dest_file = downloads_dir / filename
            with open(dest_file, "wb") as f:
                f.write(csv_content)
            saved_path_str = str(dest_file)
            try:
                os.startfile(saved_path_str)
            except Exception:
                pass
    except Exception as e:
        print(f"Errore salvataggio diretto CSV: {e}")

    return Response(
        content=csv_content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "text/csv; charset=utf-8",
            "Content-Length": str(len(csv_content)),
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "X-Saved-Path": saved_path_str,
            "Access-Control-Expose-Headers": "Content-Disposition, X-Saved-Path",
        },
    )


# ---------------------------------------------------------------------------
# GET /api/ricevute/esporta-excel — Export Excel nativo (.xlsx) per range date
# Colonne come da specifica utente: DATA | COGNOME NOME | VIA CITTA' CAP | CODICE FISCALE | PRESTAZIONE | IMPORTO
# ---------------------------------------------------------------------------
@router.get("/ricevute/esporta-excel")
@router.get("/ricevute/esporta-excel.xlsx")
def esporta_excel(
    data_inizio: Optional[date] = Query(None, description="Data inizio intervallo (inclusa)"),
    data_fine: Optional[date] = Query(None, description="Data fine intervallo (inclusa)"),
    db: Session = Depends(get_db),
):
    """
    Genera un file Excel nativo (.xlsx) con il riepilogo contabile delle prestazioni
    emesse nel periodo specificato, con le colonne richieste dalla contabilità:
    DATA | COGNOME NOME | VIA CITTA' CAP | CODICE FISCALE | PRESTAZIONE | IMPORTO
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    query = (
        db.query(Ricevuta)
        .options(joinedload(Ricevuta.paziente), joinedload(Ricevuta.righe))
    )

    if data_inizio is not None:
        query = query.filter(Ricevuta.data_emissione >= data_inizio)
    if data_fine is not None:
        query = query.filter(Ricevuta.data_emissione <= data_fine)

    ricevute = query.order_by(Ricevuta.data_emissione.asc(), Ricevuta.numero.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prestazioni Sanitarie"

    # Abilita esplicitamente le linee guida della griglia
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True

    headers = [
        "DATA",
        "COGNOME NOME",
        "VIA CITTA' CAP",
        "CODICE FISCALE",
        "PRESTAZIONE",
        "IMPORTO",
    ]
    ws.append(headers)

    # Stile riga di intestazione conforme alla foto (sfondo bianco/trasparente, testo in grassetto nero)
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    data_font = Font(name="Calibri", size=11)

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        if col_idx == 6:  # IMPORTO a destra
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")

    # Inserimento dati dal DB
    for r in ricevute:
        paz = r.paziente
        cognome_nome = f"{paz.cognome or ''} {paz.nome or ''}".strip().upper()

        # Composizione VIA CITTA' CAP
        parti_ind = []
        if paz.indirizzo and paz.indirizzo.strip():
            parti_ind.append(paz.indirizzo.strip())
        if paz.citta and paz.citta.strip():
            parti_ind.append(paz.citta.strip())
        if paz.cap and paz.cap.strip():
            parti_ind.append(paz.cap.strip())
        via_citta_cap = " ".join(parti_ind).upper() if parti_ind else ""

        data_str = r.data_emissione.strftime("%d/%m/%Y")
        cf_str = (paz.codice_fiscale or "").strip().upper()

        # Una riga per ciascuna prestazione
        # Tutte le prestazioni della fattura in fila (separate da virgola)
        descrizioni = [riga.descrizione.strip() for riga in r.righe if riga.descrizione and riga.descrizione.strip()]
        prestazioni_str = ", ".join(descrizioni)

        ws.append([
            data_str,
            cognome_nome,
            via_citta_cap,
            cf_str,
            prestazioni_str,
            float(r.totale_da_pagare or 0),
        ])

    # Formattazione celle dati (font Calibri 11, allineamenti e formato valuta per importi)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for idx, cell in enumerate(row):
            cell.font = data_font
            if idx == 5:  # IMPORTO
                cell.number_format = '#,##0.00 €'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Larghezza colonne con minimi confortevoli per non tagliare il testo
    min_widths = {1: 14, 2: 28, 3: 35, 4: 20, 5: 35, 6: 16}
    for col in ws.columns:
        col_idx = col[0].column
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        min_w = min_widths.get(col_idx, 14)
        ws.column_dimensions[col_letter].width = max(max_len + 4, min_w)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename_parts = ["ricevute"]
    if data_inizio:
        filename_parts.append(f"da_{data_inizio.strftime('%Y%m%d')}")
    if data_fine:
        filename_parts.append(f"a_{data_fine.strftime('%Y%m%d')}")
    filename = f"{'_'.join(filename_parts)}.xlsx"

    excel_bytes = stream.getvalue()

    saved_path_str = ""
    try:
        downloads_dir = Path.home() / "Downloads"
        if downloads_dir.exists():
            dest_file = downloads_dir / filename
            with open(dest_file, "wb") as f:
                f.write(excel_bytes)
            saved_path_str = str(dest_file)
            try:
                os.startfile(saved_path_str)
            except Exception:
                pass
    except Exception as e:
        print(f"Errore salvataggio diretto Excel: {e}")

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": str(len(excel_bytes)),
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "X-Saved-Path": saved_path_str,
            "Access-Control-Expose-Headers": "Content-Disposition, X-Saved-Path",
        },
    )


# ---------------------------------------------------------------------------
# GET /api/ricevute/{ricevuta_id} — Dettaglio singola ricevuta
# (Questa route con parametro path va DOPO tutte le route con path fissi)
# ---------------------------------------------------------------------------
@router.get("/ricevute/{ricevuta_id}", response_model=RicevutaOut)
def dettaglio_ricevuta(ricevuta_id: int, db: Session = Depends(get_db)):
    """
    Restituisce tutti i dati di una ricevuta (paziente + righe)
    per la ristampa fedele del documento A4.
    """
    ricevuta = (
        db.query(Ricevuta)
        .options(joinedload(Ricevuta.paziente), joinedload(Ricevuta.righe))
        .filter(Ricevuta.id == ricevuta_id)
        .first()
    )
    if ricevuta is None:
        raise HTTPException(status_code=404, detail="Ricevuta non trovata")
    return ricevuta


# ---------------------------------------------------------------------------
# GET /api/pazienti/cerca — Cerca paziente per Codice Fiscale
# ---------------------------------------------------------------------------
@router.get("/pazienti/cerca", response_model=Optional[PazienteOut])
def cerca_paziente(
    cf: str = Query(..., min_length=1, description="Codice Fiscale (anche parziale)"),
    db: Session = Depends(get_db),
):
    """
    Cerca un paziente per Codice Fiscale esatto.
    Utile per l'autocomplete nel form: se il CF esiste, precompila i campi.
    """
    paziente = (
        db.query(Paziente)
        .filter(Paziente.codice_fiscale == cf.upper().strip())
        .first()
    )
    if paziente is None:
        raise HTTPException(status_code=404, detail="Paziente non trovato")
    return paziente


# ---------------------------------------------------------------------------
# PUT /api/ricevute/{ricevuta_id} — Modifica ricevuta esistente
# ---------------------------------------------------------------------------
@router.put("/ricevute/{ricevuta_id}", response_model=RicevutaOut)
def modifica_ricevuta(
    ricevuta_id: int,
    payload: RicevutaIn,
    db: Session = Depends(get_db),
):
    """
    Modifica una ricevuta esistente (dati paziente, prestazioni, bollo, totali, data).
    Mantiene il numero progressivo e l'anno originali.
    """
    ricevuta = (
        db.query(Ricevuta)
        .options(joinedload(Ricevuta.paziente), joinedload(Ricevuta.righe))
        .filter(Ricevuta.id == ricevuta_id)
        .first()
    )
    if ricevuta is None:
        raise HTTPException(status_code=404, detail="Ricevuta non trovata")

    # 1. Aggiorna o crea il paziente
    paziente = (
        db.query(Paziente)
        .filter(Paziente.codice_fiscale == payload.paziente.codice_fiscale)
        .first()
    )
    if paziente is None:
        paziente = Paziente(**payload.paziente.model_dump())
        db.add(paziente)
        db.flush()
    else:
        for campo, valore in payload.paziente.model_dump().items():
            setattr(paziente, campo, valore)
        db.flush()

    ricevuta.paziente_id = paziente.id

    # 2. Ricalcola i totali
    subtotale = sum(riga.importo for riga in payload.righe)
    bollo_applicato = payload.bollo_applicato
    importo_bollo = Decimal("2.00") if bollo_applicato else Decimal("0.00")

    totale_da_pagare = subtotale
    if bollo_applicato and payload.bollo_a_carico_paziente:
        totale_da_pagare += importo_bollo

    # 3. Aggiorna campi testata e gestisce eventuale cambio numero
    nuovo_anno = payload.data_emissione.year
    if payload.numero is not None:
        if payload.numero != ricevuta.numero or nuovo_anno != ricevuta.anno:
            esistente = (
                db.query(Ricevuta)
                .filter(
                    Ricevuta.anno == nuovo_anno,
                    Ricevuta.numero == payload.numero,
                    Ricevuta.id != ricevuta.id,
                )
                .first()
            )
            if esistente is not None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Il numero {payload.numero} per l'anno {nuovo_anno} è già utilizzato da un'altra ricevuta",
                )
            ricevuta.numero = payload.numero

    ricevuta.anno = nuovo_anno
    ricevuta.data_emissione = payload.data_emissione
    ricevuta.subtotale = subtotale
    ricevuta.bollo_applicato = bollo_applicato
    ricevuta.bollo_a_carico_paziente = payload.bollo_a_carico_paziente
    ricevuta.importo_bollo = importo_bollo
    ricevuta.totale_da_pagare = totale_da_pagare
    ricevuta.metodo_pagamento = payload.metodo_pagamento
    ricevuta.note = payload.note

    # 4. Sostituisce le righe prestazione
    db.query(RigaRicevuta).filter(RigaRicevuta.ricevuta_id == ricevuta.id).delete()
    db.flush()

    for riga_in in payload.righe:
        riga = RigaRicevuta(
            ricevuta_id=ricevuta.id,
            ordine=riga_in.ordine,
            descrizione=riga_in.descrizione,
            importo=riga_in.importo,
        )
        db.add(riga)

    db.commit()
    db.refresh(ricevuta)

    return ricevuta


# ---------------------------------------------------------------------------
# DELETE /api/ricevute/{ricevuta_id} — Elimina ricevuta
# ---------------------------------------------------------------------------
@router.delete("/ricevute/{ricevuta_id}")
def elimina_ricevuta(ricevuta_id: int, db: Session = Depends(get_db)):
    """
    Elimina una ricevuta dal database (e le relative righe a cascata).
    """
    ricevuta = db.query(Ricevuta).filter(Ricevuta.id == ricevuta_id).first()
    if ricevuta is None:
        raise HTTPException(status_code=404, detail="Ricevuta non trovata")

    descrizione = f"Ricevuta n. {ricevuta.anno}/{ricevuta.numero}"
    db.delete(ricevuta)
    db.commit()

    return {"message": f"{descrizione} eliminata con successo"}


